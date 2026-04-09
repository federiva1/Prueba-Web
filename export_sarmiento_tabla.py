import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEAM_ID = 10086
OUTPUT_FILE = r"C:\Users\feder\OneDrive\Escritorio\argentinos_vs_sarmiento_tabla.xlsx"

with open("next_data.json", encoding="utf-8") as f:
    nd = json.load(f)

ps = nd["props"]["pageProps"]["content"]["playerStats"]

# Secciones y columnas en orden FotMob (sin Shotmap)
SECTIONS = [
    ("Top Stats", "top_stats", [
        "FotMob rating", "Minutes played", "Goals", "Assists",
        "Expected goals (xG)", "Expected goals on target (xGOT)",
        "Expected assists (xA)", "xG + xA",
        "Accurate passes", "Chances created",
        "Shots on target", "Shots off target", "Shot accuracy",
        "Defensive actions",
    ]),
    ("Ataque", "attack", [
        "Touches", "Touches in opposition box", "Successful dribbles",
        "Passes into final third", "Accurate crosses", "Corners",
        "Dispossessed", "xG Non-penalty",
    ]),
    ("Defensa", "defense", [
        "Defensive actions", "Tackles", "Blocks", "Clearances",
        "Interceptions", "Recoveries", "Dribbled past",
    ]),
    ("Duelos", "duels", [
        "Ground duels won", "Aerial duels won",
        "Was fouled", "Fouls committed",
        "Duels won", "Duels lost",
    ]),
]

# Extraer jugadores de AJ
players = []
for pid, p in ps.items():
    if str(p.get("teamId", "")) != str(TEAM_ID):
        continue
    player_stats = {}
    for section in p.get("stats", []):
        for stat_name, stat_obj in section.get("stats", {}).items():
            if stat_name == "Shotmap":
                continue
            stat = stat_obj.get("stat", {})
            value = stat.get("value", "")
            total = stat.get("total", None)
            player_stats[stat_name] = f"{value}/{total}" if total is not None else value
    players.append({
        "name": p.get("name", ""),
        "stats": player_stats,
        "minutes": player_stats.get("Minutes played", 0) or 0,
    })

# Ordenar: primero los que jugaron (por minutos desc), luego los que no
players.sort(key=lambda x: (-float(x["minutes"]) if x["minutes"] else 0, x["name"]))

# --- Construir Excel ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "vs Sarmiento"

# Colores por sección
COLORS = {
    "Top Stats": "1F4E79",
    "Ataque":    "833C00",
    "Defensa":   "375623",
    "Duelos":    "4B2B6B",
}
HEADER_FONT_COLOR = "FFFFFF"
SUBHEADER_COLORS = {
    "Top Stats": "2E75B6",
    "Ataque":    "C55A11",
    "Defensa":   "548235",
    "Duelos":    "7030A0",
}
ROW_ALT = "F2F2F2"

thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
section_border_right = Border(left=thin, right=Side(style="medium", color="888888"), top=thin, bottom=thin)

# Fila 1: columna Jugador + encabezados de sección (merged)
# Fila 2: sub-encabezados de cada stat
# Filas 3+: datos

# Calcular posiciones de columnas
col_start = 2  # columna B en adelante (A = Jugador)
section_ranges = []
for sec_name, sec_key, cols in SECTIONS:
    section_ranges.append((sec_name, col_start, col_start + len(cols) - 1, cols))
    col_start += len(cols)

total_cols = col_start - 1

# --- Fila 1: encabezado "Jugador" + secciones ---
ws.cell(1, 1, "Jugador").font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
ws.cell(1, 1).fill = PatternFill("solid", fgColor="203864")
ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
ws.cell(1, 1).border = border

for sec_name, c_start, c_end, _ in section_ranges:
    color = COLORS[sec_name]
    if c_start == c_end:
        cell = ws.cell(1, c_start)
        cell.value = sec_name
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    else:
        ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
        cell = ws.cell(1, c_start)
        cell.value = sec_name
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        for c in range(c_start, c_end + 1):
            ws.cell(1, c).fill = PatternFill("solid", fgColor=color)

# --- Fila 2: sub-encabezados de cada stat ---
ws.cell(2, 1, "Jugador").font = Font(bold=True, color=HEADER_FONT_COLOR)
ws.cell(2, 1).fill = PatternFill("solid", fgColor="203864")
ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.cell(2, 1).border = border

for sec_name, c_start, c_end, cols in section_ranges:
    color = SUBHEADER_COLORS[sec_name]
    for j, col_name in enumerate(cols):
        c = c_start + j
        cell = ws.cell(2, c)
        cell.value = col_name
        cell.fill = PatternFill("solid", fgColor=color)
        cell.font = Font(bold=True, color=HEADER_FONT_COLOR, size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border if c != c_end else section_border_right

# --- Filas de datos ---
for row_idx, player in enumerate(players):
    r = row_idx + 3
    bg = "FFFFFF" if row_idx % 2 == 0 else ROW_ALT

    # Nombre
    cell = ws.cell(r, 1, player["name"])
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(vertical="center")
    cell.border = border

    for sec_name, c_start, c_end, cols in section_ranges:
        for j, col_name in enumerate(cols):
            c = c_start + j
            val = player["stats"].get(col_name, "")
            cell = ws.cell(r, c)
            # Intentar convertir a número si es posible
            if val != "" and "/" not in str(val):
                try:
                    val = float(val) if "." in str(val) else int(val)
                except (ValueError, TypeError):
                    pass
            cell.value = val
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border if c != c_end else section_border_right

# --- Anchos de columna ---
ws.column_dimensions["A"].width = 22
for sec_name, c_start, c_end, cols in section_ranges:
    for j, col_name in enumerate(cols):
        c = c_start + j
        col_letter = get_column_letter(c)
        ws.column_dimensions[col_letter].width = max(10, min(18, len(col_name) * 0.9))

ws.row_dimensions[1].height = 22
ws.row_dimensions[2].height = 38
for r in range(3, 3 + len(players)):
    ws.row_dimensions[r].height = 18

# Fijar las dos primeras filas y la columna A
ws.freeze_panes = "B3"

wb.save(OUTPUT_FILE)
print(f"Excel guardado: {OUTPUT_FILE}")
print(f"Jugadores: {len(players)} | Secciones: {len(SECTIONS)} | Stats totales: {sum(len(c) for _,_,c in SECTIONS)}")
