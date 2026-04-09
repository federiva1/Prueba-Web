# export_partido.py
# Lee el JSON de un partido (generado por scrape_partido.py) y exporta
# un Excel con la tabla "Estad. Jugador" de FotMob, agrupada por secciones.
#
# Uso:
#   python export_partido.py <next_data.json> <fecha> <rival> <Local|Visitante> <resultado> <salida.xlsx>
#
# Ejemplo:
#   python export_partido.py "nd_sarmiento.json" "2026-01-26" "Sarmiento" "Local" "1 - 0" "AJ_vs_Sarmiento.xlsx"

import sys
import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config_equipo import TEAM_ID, TEAM_NAME
from keys_fotmob import KEY_MAP, SECTIONS

if len(sys.argv) != 7:
    print("Uso: python export_partido.py <next_data.json> <fecha> <rival> <Local|Visitante> <resultado> <salida.xlsx>")
    sys.exit(1)

nd_file, fecha, rival, local_visit, resultado, output_file = sys.argv[1:]

with open(nd_file, encoding="utf-8") as f:
    nd = json.load(f)

ps = nd["props"]["pageProps"]["content"]["playerStats"]

# ── Extraer jugadores del equipo ──────────────────────────────────────────────
players = []
for pid, p in ps.items():
    if str(p.get("teamId", "")) != str(TEAM_ID):
        continue
    flat = {}
    for section in p.get("stats", []):
        for stat_name, stat_obj in section.get("stats", {}).items():
            if stat_name == "Shotmap":
                continue
            stat = stat_obj.get("stat", {})
            if stat_name not in flat:
                flat[stat_name] = (stat.get("value", None), stat.get("total", None))
    players.append({
        "name":    p.get("name", ""),
        "is_gk":  p.get("isGoalkeeper", False),
        "minutes": flat.get("Minutes played", (0, None))[0] or 0,
        "flat":    flat,
    })

players.sort(key=lambda x: (-float(x["minutes"]) if x["minutes"] else 0, x["name"]))

# ── Helpers de estilo ─────────────────────────────────────────────────────────
def fill(h):
    return PatternFill("solid", fgColor=h)

thin_s = Side(style="thin",   color="D0D0D0")
med_s  = Side(style="medium", color="777777")

def bord(rh=False):
    return Border(left=thin_s, right=med_s if rh else thin_s, top=thin_s, bottom=thin_s)

def fmt(value, total):
    if value is None:
        return ""
    if total is not None:
        return f"{value}/{total}"
    try:
        f = float(value)
        return int(f) if f == int(f) else round(f, 2)
    except (TypeError, ValueError):
        return value

# ── Calcular posición de cada sección ────────────────────────────────────────
section_info = []
col_cursor = 2
for sec_name, dark, mid, cols in SECTIONS:
    c_start = col_cursor
    c_end   = col_cursor + len(cols) - 1
    section_info.append((sec_name, dark, mid, cols, c_start, c_end))
    col_cursor = c_end + 1
total_cols = col_cursor - 1

# ── Crear workbook ────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = f"vs {rival[:28]}"

# Fila 1 — título del partido
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
t = ws.cell(1, 1)
t.value     = f"{TEAM_NAME} vs {rival}  |  {fecha}  |  {local_visit}  |  {resultado}"
t.fill      = fill("1A1A2E")
t.font      = Font(bold=True, color="FFFFFF", size=12)
t.alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, total_cols + 1):
    ws.cell(1, c).fill = fill("1A1A2E")
ws.row_dimensions[1].height = 22

# Fila 2 — encabezados de sección (mergeados)
ws.cell(2, 1, "Jugador").fill      = fill("203864")
ws.cell(2, 1).font                 = Font(bold=True, color="FFFFFF", size=11)
ws.cell(2, 1).alignment            = Alignment(horizontal="center", vertical="center")
ws.cell(2, 1).border               = bord()

for sec_name, dark, mid, cols, c_start, c_end in section_info:
    ws.merge_cells(start_row=2, start_column=c_start, end_row=2, end_column=c_end)
    cell = ws.cell(2, c_start)
    cell.value = sec_name; cell.fill = fill(dark)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = bord(rh=True)
    for c in range(c_start + 1, c_end + 1):
        ws.cell(2, c).fill = fill(dark)
        ws.cell(2, c).border = bord(rh=(c == c_end))
ws.row_dimensions[2].height = 24

# Fila 3 — sub-encabezados de cada stat
ws.cell(3, 1, "Jugador").fill      = fill("2B4DA8")
ws.cell(3, 1).font                 = Font(bold=True, color="FFFFFF", size=10)
ws.cell(3, 1).alignment            = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.cell(3, 1).border               = bord()

for sec_name, dark, mid, cols, c_start, c_end in section_info:
    for j, col_name in enumerate(cols):
        c    = c_start + j
        cell = ws.cell(3, c)
        cell.value = col_name; cell.fill = fill(mid)
        cell.font = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = bord(rh=(c == c_end))
ws.row_dimensions[3].height = 40

# Filas de datos
row_colors = ["FFFFFF", "EEF2F8"]
for row_idx, player in enumerate(players):
    r  = row_idx + 4
    bg = row_colors[row_idx % 2]

    nc = ws.cell(r, 1, player["name"])
    nc.fill = fill(bg); nc.font = Font(bold=True, size=10)
    nc.alignment = Alignment(vertical="center"); nc.border = bord()

    for sec_name, dark, mid, cols, c_start, c_end in section_info:
        for j, col_name in enumerate(cols):
            c       = c_start + j
            api_key = KEY_MAP.get(col_name, col_name)
            value, total = player["flat"].get(api_key, (None, None))
            dc = ws.cell(r, c)
            dc.value     = fmt(value, total)
            dc.fill      = fill(bg)
            dc.font      = Font(size=10)
            dc.alignment = Alignment(horizontal="center", vertical="center")
            dc.border    = bord(rh=(c == c_end))

    ws.row_dimensions[r].height = 18

# Anchos de columna
ws.column_dimensions["A"].width = 24
for sec_name, dark, mid, cols, c_start, c_end in section_info:
    for j, col_name in enumerate(cols):
        ws.column_dimensions[get_column_letter(c_start + j)].width = max(10, min(18, len(col_name) * 0.92))

ws.freeze_panes = "B4"

wb.save(output_file)
print(f"OK — Excel guardado: {output_file} | Jugadores: {len(players)} | Columnas: {total_cols}")
