import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEAM_ID = 10086
OUTPUT_FILE = r"C:\Users\feder\OneDrive\Escritorio\AJ_vs_Sarmiento_Unificada.xlsx"

with open("next_data.json", encoding="utf-8") as f:
    nd = json.load(f)

ps = nd["props"]["pageProps"]["content"]["playerStats"]

KEY_MAP = {
    "Minutos jugados":               "Minutes played",
    "Goles":                         "Goals",
    "Asistencias":                   "Assists",
    "xG":                            "Expected goals (xG)",
    "xA":                            "Expected assists (xA)",
    "xG + xA":                       "xG + xA",
    "Acciones defensivas":           "Defensive actions",
    "xGOT":                          "Expected goals on target (xGOT)",
    "Disparos a puerta":             "Shots on target",
    "Regates realizados":            "Successful dribbles",
    "Grandes oportunidades perdidas":"Big chances missed",
    "Toques en el área rival":       "Touches in opposition box",
    "Fueras de juego":               "Offsides",
    "Toques":                        "Touches",
    "Pases precisos":                "Accurate passes",
    "Oportunidades creadas":         "Chances created",
    "Pases en el último tercio":     "Passes into final third",
    "Centros precisos":              "Accurate crosses",
    "Tiros largos precisos":         "Accurate long balls",
    "Entradas":                      "Tackles",
    "Interceptaciones":              "Interceptions",
    "Bloqueos":                      "Blocks",
    "Recuperaciones":                "Recoveries",
    "Despejes":                      "Clearances",
    "Despeje de cabeza":             "Headed clearance",
    "Regateado":                     "Dribbled past",
    "Duelos ganados":                "Duels won",
    "Duelos perdidos":               "Duels lost",
    "Duelo terrestre ganado":        "Ground duels won",
    "Aéreo ganado":                  "Aerial duels won",
    "Faltas":                        "Fouls committed",
    "Faltas recibidas":              "Was fouled",
    "Paradas":                       "Saves",
    "Goles en contra":               "Goals conceded",
    "xGOT recibidos":                "xGOT faced",
    "Goles evitados":                "Goals prevented",
    "El portero actuó como líbero":  "Acted as sweeper",
    "Salida por alto":               "High claim",
}

SECTIONS = [
    ("Top Estadísticas", "1F4E79", "2E75B6", [
        "Minutos jugados", "Goles", "Asistencias",
        "xG", "xA", "xG + xA", "Acciones defensivas",
    ]),
    ("Ataque", "833C00", "C55A11", [
        "Goles", "xG", "xGOT", "Disparos a puerta",
        "Regates realizados", "Grandes oportunidades perdidas",
        "Toques en el área rival", "Fueras de juego",
    ]),
    ("Pases", "1F6B3B", "2E8B57", [
        "Toques", "Pases precisos", "Asistencias", "xA",
        "Oportunidades creadas", "Pases en el último tercio",
        "Centros precisos", "Tiros largos precisos",
    ]),
    ("Defensa", "375623", "548235", [
        "Acciones defensivas", "Entradas", "Interceptaciones", "Bloqueos",
        "Recuperaciones", "Despejes", "Despeje de cabeza", "Regateado",
    ]),
    ("Duelos", "4B2B6B", "7030A0", [
        "Duelos ganados", "Duelos perdidos", "Duelo terrestre ganado",
        "Aéreo ganado", "Faltas", "Faltas recibidas",
        "Regates realizados", "Entradas",
    ]),
    ("Portero", "1C3A5C", "2F5496", [
        "Paradas", "Goles en contra", "xGOT recibidos", "Goles evitados",
        "El portero actuó como líbero", "Salida por alto",
        "Tiros largos precisos", "Pases precisos",
    ]),
]

# ── Extraer jugadores ──
players = []
for pid, p in ps.items():
    if str(p.get("teamId", "")) != str(TEAM_ID):
        continue
    flat = {}
    for section in p.get("stats", []):
        for stat_name, stat_obj in section.get("stats", {}).items():
            if stat_name == "Shotmap":
                continue
            stat = stat_obj.get("stat", {})
            if stat_name not in flat:
                flat[stat_name] = (stat.get("value", None), stat.get("total", None))
    players.append({
        "name": p.get("name", ""),
        "is_gk": p.get("isGoalkeeper", False),
        "minutes": flat.get("Minutes played", (0, None))[0] or 0,
        "flat": flat,
    })
players.sort(key=lambda x: (-float(x["minutes"]) if x["minutes"] else 0, x["name"]))

# ── Helpers ──
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def lighten(h, f=0.52):
    r = int(int(h[0:2],16) + (255-int(h[0:2],16))*f)
    g = int(int(h[2:4],16) + (255-int(h[2:4],16))*f)
    b = int(int(h[4:6],16) + (255-int(h[4:6],16))*f)
    return f"{r:02X}{g:02X}{b:02X}"

thin_s = Side(style="thin",   color="D0D0D0")
med_s  = Side(style="medium", color="777777")

def bord(right_heavy=False):
    return Border(left=thin_s, right=med_s if right_heavy else thin_s,
                  top=thin_s, bottom=thin_s)

def fmt_val(value, total):
    if value is None:
        return ""
    if total is not None:
        return f"{value}/{total}"
    try:
        f = float(value)
        return int(f) if f == int(f) else round(f, 2)
    except (TypeError, ValueError):
        return value

# ── Calcular posición de cada sección ──
section_info = []
col_cursor = 2   # columna A = Jugador
for sec_name, dark, mid, cols in SECTIONS:
    c_start = col_cursor
    c_end   = col_cursor + len(cols) - 1
    section_info.append((sec_name, dark, mid, cols, c_start, c_end))
    col_cursor = c_end + 1
total_cols = col_cursor - 1

# ── Crear workbook ──
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Estad. Jugador"

# ── Fila 1: encabezado "Jugador" + secciones mergeadas ──
ws.cell(1, 1, "Jugador").fill     = fill("203864")
ws.cell(1, 1).font                = Font(bold=True, color="FFFFFF", size=11)
ws.cell(1, 1).alignment           = Alignment(horizontal="center", vertical="center")
ws.cell(1, 1).border              = bord()

for sec_name, dark, mid, cols, c_start, c_end in section_info:
    ws.merge_cells(start_row=1, start_column=c_start, end_row=1, end_column=c_end)
    cell = ws.cell(1, c_start)
    cell.value     = sec_name
    cell.fill      = fill(dark)
    cell.font      = Font(bold=True, color="FFFFFF", size=11)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border    = bord(right_heavy=True)
    # rellenar celdas mergeadas
    for c in range(c_start + 1, c_end + 1):
        ws.cell(1, c).fill   = fill(dark)
        ws.cell(1, c).border = bord(right_heavy=(c == c_end))
ws.row_dimensions[1].height = 24

# ── Fila 2: sub-encabezados ──
ws.cell(2, 1, "Jugador").fill      = fill("2B4DA8")
ws.cell(2, 1).font                 = Font(bold=True, color="FFFFFF", size=10)
ws.cell(2, 1).alignment            = Alignment(horizontal="center", vertical="center", wrap_text=True)
ws.cell(2, 1).border               = bord()

for sec_name, dark, mid, cols, c_start, c_end in section_info:
    for j, col_name in enumerate(cols):
        c    = c_start + j
        cell = ws.cell(2, c)
        cell.value     = col_name
        cell.fill      = fill(mid)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = bord(right_heavy=(c == c_end))
ws.row_dimensions[2].height = 40

# ── Filas de datos ──
row_colors = ["FFFFFF", "EEF2F8"]
for row_idx, player in enumerate(players):
    r  = row_idx + 3
    bg = row_colors[row_idx % 2]

    cell = ws.cell(r, 1, player["name"])
    cell.fill      = fill(bg)
    cell.font      = Font(bold=True, size=10)
    cell.alignment = Alignment(vertical="center")
    cell.border    = bord()

    for sec_name, dark, mid, cols, c_start, c_end in section_info:
        for j, col_name in enumerate(cols):
            c       = c_start + j
            api_key = KEY_MAP.get(col_name, col_name)
            value, total = player["flat"].get(api_key, (None, None))
            dc = ws.cell(r, c)
            dc.value     = fmt_val(value, total)
            dc.fill      = fill(bg)
            dc.font      = Font(size=10)
            dc.alignment = Alignment(horizontal="center", vertical="center")
            dc.border    = bord(right_heavy=(c == c_end))

    ws.row_dimensions[r].height = 18

# ── Anchos de columna ──
ws.column_dimensions["A"].width = 24
for sec_name, dark, mid, cols, c_start, c_end in section_info:
    for j, col_name in enumerate(cols):
        col_letter = get_column_letter(c_start + j)
        ws.column_dimensions[col_letter].width = max(10, min(18, len(col_name) * 0.92))

ws.freeze_panes = "B3"

wb.save(OUTPUT_FILE)
print(f"Excel guardado: {OUTPUT_FILE}")
print(f"Jugadores: {len(players)} | Secciones: {len(SECTIONS)} | Columnas totales: {total_cols}")
