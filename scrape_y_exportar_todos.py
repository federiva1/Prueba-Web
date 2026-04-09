# scrape_y_exportar_todos.py
# Scrapea todos los partidos de una lista y los exporta en un único Excel,
# con una hoja por partido, manteniendo el formato FotMob "Estad. Jugador".

import json
import time
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright
from config_equipo import TEAM_ID, TEAM_NAME
from keys_fotmob import KEY_MAP, SECTIONS

OUTPUT_FILE = r"C:\Users\feder\OneDrive\Escritorio\AJ_Apertura2026_Todos.xlsx"

# ── Partidos a procesar ───────────────────────────────────────────────────────
PARTIDOS = [
    {"fecha": "2026-02-03", "rival": "Belgrano",               "local": True,  "resultado": "0 - 0", "url": "https://www.fotmob.com/matches/argentinos-juniors-vs-belgrano/3d7rp3#5101909"},
    {"fecha": "2026-02-07", "rival": "Racing Club",            "local": False, "resultado": "2 - 1", "url": "https://www.fotmob.com/matches/racing-club-vs-argentinos-juniors/3d2kwr#5101935"},
    {"fecha": "2026-02-13", "rival": "River Plate",            "local": False, "resultado": "1 - 0", "url": "https://www.fotmob.com/matches/river-plate-vs-argentinos-juniors/3d0uo9#5101939"},
    {"fecha": "2026-03-01", "rival": "Barracas Central",       "local": True,  "resultado": "1 - 1", "url": "https://www.fotmob.com/matches/argentinos-juniors-vs-barracas-central/bhib0fc#5101984"},
    {"fecha": "2026-03-11", "rival": "Rosario Central",        "local": False, "resultado": "0 - 0", "url": "https://www.fotmob.com/matches/rosario-central-vs-argentinos-juniors/3d4b5p#5102014"},
    {"fecha": "2026-03-16", "rival": "Tigre",                  "local": False, "resultado": "1 - 1", "url": "https://www.fotmob.com/matches/argentinos-juniors-vs-tigre/29u6yfn#5102042"},
    {"fecha": "2026-03-23", "rival": "Club Atletico Platense", "local": True,  "resultado": "1 - 0", "url": "https://www.fotmob.com/matches/argentinos-juniors-vs-club-atletico-platense/3d6gzl#5102044"},
    {"fecha": "2026-03-27", "rival": "Lanus",                  "local": True,  "resultado": "2 - 1", "url": "https://www.fotmob.com/matches/lanus-vs-argentinos-juniors/3d3g16#5101954"},
    {"fecha": "2026-03-31", "rival": "Aldosivi",               "local": False, "resultado": "0 - 2", "url": "https://www.fotmob.com/matches/argentinos-juniors-vs-aldosivi/6s3vvhh#5101969"},
    {"fecha": "2026-04-06", "rival": "Banfield",               "local": True,  "resultado": "3 - 2", "url": "https://www.fotmob.com/matches/argentinos-juniors-vs-banfield/3d5luq#5102060"},
]

# ── Helpers de estilo ─────────────────────────────────────────────────────────
def fill(h):    return PatternFill("solid", fgColor=h)
thin_s = Side(style="thin",   color="D0D0D0")
med_s  = Side(style="medium", color="777777")
def bord(rh=False): return Border(left=thin_s, right=med_s if rh else thin_s, top=thin_s, bottom=thin_s)
def fmt(value, total):
    if value is None: return ""
    if total is not None: return f"{value}/{total}"
    try:
        f = float(value); return int(f) if f == int(f) else round(f, 2)
    except: return value

# ── Calcular posición de columnas (fijo para todas las hojas) ─────────────────
section_info = []
col_cursor = 2
for sec_name, dark, mid, cols in SECTIONS:
    section_info.append((sec_name, dark, mid, cols, col_cursor, col_cursor + len(cols) - 1))
    col_cursor += len(cols)
total_cols = col_cursor - 1

# ── Extraer jugadores desde __NEXT_DATA__ ─────────────────────────────────────
def extraer_jugadores(nd):
    ps = nd["props"]["pageProps"]["content"]["playerStats"]
    players = []
    for pid, p in ps.items():
        if str(p.get("teamId", "")) != str(TEAM_ID):
            continue
        flat = {}
        for section in p.get("stats", []):
            for stat_name, stat_obj in section.get("stats", {}).items():
                if stat_name == "Shotmap": continue
                stat = stat_obj.get("stat", {})
                if stat_name not in flat:
                    flat[stat_name] = (stat.get("value", None), stat.get("total", None))
        players.append({
            "name":    p.get("name", ""),
            "is_gk":  p.get("isGoalkeeper", False),
            "minutes": flat.get("Minutes played", (0, None))[0] or 0,
            "flat":    flat,
        })
    players.sort(key=lambda x: (-float(x["minutes"]) if x["minutes"] else 0, x["name"]))
    return players

# ── Escribir una hoja en el workbook ─────────────────────────────────────────
def escribir_hoja(wb, partido, players):
    fecha      = partido["fecha"]
    rival      = partido["rival"]
    local_visit = "Local" if partido["local"] else "Visitante"
    resultado  = partido["resultado"]
    sheet_name = f"{fecha[5:]}  {rival[:18]}"   # ej: "02-03  Belgrano"

    ws = wb.create_sheet(title=sheet_name)
    row_colors = ["FFFFFF", "EEF2F8"]

    # Fila 1 — título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    t = ws.cell(1, 1)
    t.value     = f"{TEAM_NAME} vs {rival}  |  {fecha}  |  {local_visit}  |  {resultado}"
    t.fill      = fill("1A1A2E")
    t.font      = Font(bold=True, color="FFFFFF", size=12)
    t.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(2, total_cols + 1):
        ws.cell(1, c).fill = fill("1A1A2E")
    ws.row_dimensions[1].height = 22

    # Fila 2 — secciones
    ws.cell(2, 1, "Jugador").fill = fill("203864")
    ws.cell(2, 1).font = Font(bold=True, color="FFFFFF", size=11)
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(2, 1).border = bord()
    for sec_name, dark, mid, cols, c_start, c_end in section_info:
        ws.merge_cells(start_row=2, start_column=c_start, end_row=2, end_column=c_end)
        cell = ws.cell(2, c_start)
        cell.value = sec_name; cell.fill = fill(dark)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bord(rh=True)
        for c in range(c_start + 1, c_end + 1):
            ws.cell(2, c).fill = fill(dark)
            ws.cell(2, c).border = bord(rh=(c == c_end))
    ws.row_dimensions[2].height = 24

    # Fila 3 — sub-encabezados
    ws.cell(3, 1, "Jugador").fill = fill("2B4DA8")
    ws.cell(3, 1).font = Font(bold=True, color="FFFFFF", size=10)
    ws.cell(3, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(3, 1).border = bord()
    for sec_name, dark, mid, cols, c_start, c_end in section_info:
        for j, col_name in enumerate(cols):
            c = c_start + j
            cell = ws.cell(3, c)
            cell.value = col_name; cell.fill = fill(mid)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = bord(rh=(c == c_end))
    ws.row_dimensions[3].height = 40

    # Filas de datos
    for row_idx, player in enumerate(players):
        r = row_idx + 4; bg = row_colors[row_idx % 2]
        nc = ws.cell(r, 1, player["name"])
        nc.fill = fill(bg); nc.font = Font(bold=True, size=10)
        nc.alignment = Alignment(vertical="center"); nc.border = bord()
        for sec_name, dark, mid, cols, c_start, c_end in section_info:
            for j, col_name in enumerate(cols):
                c = c_start + j
                api_key = KEY_MAP.get(col_name, col_name)
                value, total = player["flat"].get(api_key, (None, None))
                dc = ws.cell(r, c)
                dc.value = fmt(value, total); dc.fill = fill(bg)
                dc.font = Font(size=10)
                dc.alignment = Alignment(horizontal="center", vertical="center")
                dc.border = bord(rh=(c == c_end))
        ws.row_dimensions[r].height = 18

    # Anchos
    ws.column_dimensions["A"].width = 24
    for sec_name, dark, mid, cols, c_start, c_end in section_info:
        for j, col_name in enumerate(cols):
            ws.column_dimensions[get_column_letter(c_start + j)].width = max(10, min(18, len(col_name) * 0.92))
    ws.freeze_panes = "B4"

# ── Main: scrapear y exportar ─────────────────────────────────────────────────
wb = openpyxl.Workbook()
wb.remove(wb.active)

with sync_playwright() as p:
    browser = p.chromium.launch(headless=False, slow_mo=100)
    context = browser.new_context(
        user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
        locale="es-AR",
        viewport={"width": 1280, "height": 900}
    )
    page = context.new_page()

    for i, partido in enumerate(PARTIDOS):
        rival = partido["rival"]
        fecha = partido["fecha"]
        print(f"[{i+1}/{len(PARTIDOS)}] {fecha} vs {rival} ...", end=" ", flush=True)

        try:
            page.goto(partido["url"], wait_until="domcontentloaded", timeout=45000)
            time.sleep(5)
            nd = page.evaluate("() => window.__NEXT_DATA__")
            if not nd:
                print("ERROR: sin __NEXT_DATA__")
                continue
            players = extraer_jugadores(nd)
            escribir_hoja(wb, partido, players)
            print(f"OK ({len(players)} jugadores)")
        except Exception as e:
            print(f"ERROR: {e}")

        time.sleep(2)

    browser.close()

wb.save(OUTPUT_FILE)
print(f"\nExcel guardado: {OUTPUT_FILE}")
print(f"Hojas: {len(wb.sheetnames)}")
