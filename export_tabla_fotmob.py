import json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TEAM_ID = 10086
OUTPUT_FILE = r"C:\Users\feder\OneDrive\Escritorio\AJ_vs_Sarmiento_EstadjJugador.xlsx"

with open("next_data.json", encoding="utf-8") as f:
    nd = json.load(f)

ps = nd["props"]["pageProps"]["content"]["playerStats"]

# ── Mapeo: nombre en español → key en la API (puede estar en cualquier sección) ──
KEY_MAP = {
    "Minutos jugados":               "Minutes played",
    "Goles":                         "Goals",
    "Asistencias":                   "Assists",
    "xG":                            "Expected goals (xG)",
    "xA":                            "Expected assists (xA)",
    "xG + xA":                       "xG + xA",
    "Acciones defensivas":           "Defensive actions",
    "xGOT":                          "Expected goals on target (xGOT)",
    "Disparos a puerta":             "Shots on target",
    "Regates realizados":            "Successful dribbles",
    "Grandes oportunidades perdidas":"Big chances missed",
    "Toques en el área rival":       "Touches in opposition box",
    "Fueras de juego":               "Offsides",
    "Toques":                        "Touches",
    "Pases precisos":                "Accurate passes",
    "Oportunidades creadas":         "Chances created",
    "Pases en el último tercio":     "Passes into final third",
    "Centros precisos":              "Accurate crosses",
    "Tiros largos precisos":         "Accurate long balls",
    "Entradas":                      "Tackles",
    "Interceptaciones":              "Interceptions",
    "Bloqueos":                      "Blocks",
    "Recuperaciones":                "Recoveries",
    "Despejes":                      "Clearances",
    "Despeje de cabeza":             "Headed clearance",
    "Regateado":                     "Dribbled past",
    "Duelos ganados":                "Duels won",
    "Duelos perdidos":               "Duels lost",
    "Duelo terrestre ganado":        "Ground duels won",
    "Aéreo ganado":                  "Aerial duels won",
    "Faltas":                        "Fouls committed",
    "Faltas recibidas":              "Was fouled",
    "Paradas":                       "Saves",
    "Goles en contra":               "Goals conceded",
    "xGOT recibidos":                "xGOT faced",
    "Goles evitados":                "Goals prevented",
    "El portero actuó como líbero":  "Acted as sweeper",
    "Salida por alto":               "High claim",
}

# ── Definición de cada hoja: (nombre_hoja, color_hex, [columnas en español]) ──
SHEETS = [
    ("Top Estadísticas", "1F4E79", [
        "Minutos jugados", "Goles", "Asistencias",
        "xG", "xA", "xG + xA", "Acciones defensivas",
    ]),
    ("Ataque", "833C00", [
        "Goles", "xG", "xGOT", "Disparos a puerta",
        "Regates realizados", "Grandes oportunidades perdidas",
        "Toques en el área rival", "Fueras de juego",
    ]),
    ("Pases", "1F6B3B", [
        "Toques", "Pases precisos", "Asistencias", "xA",
        "Oportunidades creadas", "Pases en el último tercio",
        "Centros precisos", "Tiros largos precisos",
    ]),
    ("Defensa", "375623", [
        "Acciones defensivas", "Entradas", "Interceptaciones", "Bloqueos",
        "Recuperaciones", "Despejes", "Despeje de cabeza", "Regateado",
    ]),
    ("Duelos", "4B2B6B", [
        "Duelos ganados", "Duelos perdidos", "Duelo terrestre ganado",
        "Aéreo ganado", "Faltas", "Faltas recibidas",
        "Regates realizados", "Entradas",
    ]),
    ("Portero", "1C3A5C", [
        "Paradas", "Goles en contra", "xGOT recibidos", "Goles evitados",
        "El portero actuó como líbero", "Salida por alto",
        "Tiros largos precisos", "Pases precisos",
    ]),
]

# ── Extraer todos los stats de cada jugador de AJ (barriendo todas las secciones) ──
players = []
for pid, p in ps.items():
    if str(p.get("teamId", "")) != str(TEAM_ID):
        continue
    flat = {}
    for section in p.get("stats", []):
        for stat_name, stat_obj in section.get("stats", {}).items():
            if stat_name == "Shotmap":
                continue
            stat = stat_obj.get("stat", {})
            value = stat.get("value", None)
            total = stat.get("total", None)
            if stat_name not in flat:   # primera sección gana (top_stats tiene prioridad)
                flat[stat_name] = (value, total)
    players.append({
        "name": p.get("name", ""),
        "is_gk": p.get("isGoalkeeper", False),
        "minutes": flat.get("Minutes played", (0, None))[0] or 0,
        "flat": flat,
    })

players.sort(key=lambda x: (-float(x["minutes"]) if x["minutes"] else 0, x["name"]))

# ── Helpers de estilo ──
def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def lighten(hex_color, factor=0.55):
    """Versión más clara de un color hex para filas alternas."""
    r = int(hex_color[0:2], 16)
    g = int(hex_color[2:4], 16)
    b = int(hex_color[4:6], 16)
    r = int(r + (255 - r) * factor)
    g = int(g + (255 - g) * factor)
    b = int(b + (255 - b) * factor)
    return f"{r:02X}{g:02X}{b:02X}"

def header_font(size=10):
    return Font(bold=True, color="FFFFFF", size=size)

def data_font(bold=False):
    return Font(bold=bold, size=10)

thin = Side(style="thin", color="D0D0D0")
med  = Side(style="medium", color="888888")
def cell_border(right_heavy=False):
    r = med if right_heavy else thin
    return Border(left=thin, right=r, top=thin, bottom=thin)

def write_sheet(wb, sheet_name, color_hex, col_names, players, only_gk=False):
    ws = wb.create_sheet(title=sheet_name)
    light = lighten(color_hex)
    row_colors = ["FFFFFF", light]

    # ── Fila 1: encabezado de sección (merge) ──
    total_cols = 1 + len(col_names)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    hdr = ws.cell(1, 1)
    hdr.value = sheet_name.upper()
    hdr.fill = make_fill(color_hex)
    hdr.font = Font(bold=True, color="FFFFFF", size=12)
    hdr.alignment = Alignment(horizontal="center", vertical="center")
    for c in range(2, total_cols + 1):
        ws.cell(1, c).fill = make_fill(color_hex)
    ws.row_dimensions[1].height = 24

    # ── Fila 2: sub-encabezados ──
    sub_color = lighten(color_hex, 0.25)
    ws.cell(2, 1, "Jugador").fill = make_fill(sub_color)
    ws.cell(2, 1).font = header_font()
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(2, 1).border = cell_border()

    for j, col_name in enumerate(col_names):
        c = j + 2
        cell = ws.cell(2, c)
        cell.value = col_name
        cell.fill = make_fill(sub_color)
        cell.font = header_font(9)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = cell_border(right_heavy=(c == total_cols))
    ws.row_dimensions[2].height = 36

    # ── Filas de datos ──
    row_idx = 0
    for player in players:
        if only_gk and not player["is_gk"]:
            continue
        r = row_idx + 3
        bg = row_colors[row_idx % 2]

        # Nombre
        nc = ws.cell(r, 1, player["name"])
        nc.fill = make_fill(bg)
        nc.font = Font(bold=True, size=10)
        nc.alignment = Alignment(vertical="center")
        nc.border = cell_border()

        for j, col_name in enumerate(col_names):
            c = j + 2
            api_key = KEY_MAP.get(col_name, col_name)
            value, total = player["flat"].get(api_key, (None, None))

            if value is None:
                display = ""
            elif total is not None:
                display = f"{value}/{total}"
            else:
                # Convertir a int si es entero, float si tiene decimales
                try:
                    display = int(value) if float(value) == int(float(value)) else round(float(value), 2)
                except (TypeError, ValueError):
                    display = value

            dc = ws.cell(r, c)
            dc.value = display
            dc.fill = make_fill(bg)
            dc.font = data_font()
            dc.alignment = Alignment(horizontal="center", vertical="center")
            dc.border = cell_border(right_heavy=(c == total_cols))

        ws.row_dimensions[r].height = 18
        row_idx += 1

    # ── Anchos ──
    ws.column_dimensions["A"].width = 24
    for j, col_name in enumerate(col_names):
        col_letter = get_column_letter(j + 2)
        ws.column_dimensions[col_letter].width = max(11, min(20, len(col_name) * 0.95))

    ws.freeze_panes = "B3"
    return ws


# ── Generar workbook ──
wb = openpyxl.Workbook()
wb.remove(wb.active)  # quitar hoja vacía default

for sheet_name, color, cols in SHEETS:
    only_gk = (sheet_name == "Portero")
    write_sheet(wb, sheet_name, color, cols, players, only_gk=only_gk)

wb.save(OUTPUT_FILE)
print(f"Excel guardado: {OUTPUT_FILE}")
print(f"Jugadores: {len(players)} | Hojas: {len(SHEETS)}")
for p in players:
    mins = p['minutes']
    print(f"  {'[GK] ' if p['is_gk'] else '     '}{p['name']:30s} {mins} min")
